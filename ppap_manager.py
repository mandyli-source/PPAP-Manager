"""
PPAP Manager
============
Folder naming rule:  <PI>-<PN>-<CUSTOMER>
Example:  S383A014-06104173AA-1CYNAO
            PI       = S383A014    (Order number)
            PN       = 06104173AA  (Customer Part Number)
            CUSTOMER = 1CYNAO      (Customer Code)

Structure:
  data/
    <Engineer>/
      <Customer>/
        <PI>-<PN>-<Customer>/
          *.xlsx  *.docx  *.xdw

Run locally:
    streamlit run ppap_manager.py

On Streamlit Cloud:
    Upload PPAP files into the  data/  folder in your GitHub repo.
    The app will read from there automatically.
"""

import streamlit as st
import pandas as pd
import os
import subprocess
import platform
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="PPAP Manager",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────

# Auto-detect environment:
# - Running locally on Windows  → use C:\Users\...\Desktop\PPAP
# - Running on Streamlit Cloud  → use data/ folder in repo
IS_WINDOWS = platform.system() == "Windows"
DEFAULT_FOLDER = r"C:\Users\S2234009\Desktop\PPAP" if IS_WINDOWS else "data"

EXCEL_EXT = {".xlsx", ".xls", ".xlsm"}
WORD_EXT  = {".docx", ".doc"}
XDW_EXT   = {".xdw"}
ALL_EXT   = EXCEL_EXT | WORD_EXT | XDW_EXT

DOCUWORKS_PATHS = [
    r"C:\Program Files\Fuji Xerox\DocuWorks\deskew.exe",
    r"C:\Program Files (x86)\Fuji Xerox\DocuWorks\deskew.exe",
    r"C:\Program Files\Fujifilm\DocuWorks\deskew.exe",
    r"C:\Program Files (x86)\Fujifilm\DocuWorks\deskew.exe",
]

# ─────────────────────────────────────────────────────────────────────────────
# BLOCK 1 — DATA PIPELINE
# ─────────────────────────────────────────────────────────────────────────────

def parse_folder_name(name: str) -> dict:
    """
    Parse folder name: <PI>-<PN>-<CUSTOMER>
    e.g. S383A014-06104173AA-1CYNAO
         PI=S383A014  PN=06104173AA  CUSTOMER=1CYNAO
    """
    parts = name.split("-", 2)
    return {
        "PI":       parts[0].strip() if len(parts) > 0 else name,
        "PN":       parts[1].strip() if len(parts) > 1 else "—",
        "CUSTOMER": parts[2].strip() if len(parts) > 2 else "—",
    }


def classify_files(folder_path: Path) -> dict:
    """Group files directly inside folder_path by type."""
    out = {"excel": [], "word": [], "xdw": [], "other": []}
    for f in sorted(folder_path.iterdir()):
        if not f.is_file():
            continue
        ext = f.suffix.lower()
        if ext in EXCEL_EXT:
            out["excel"].append(f)
        elif ext in WORD_EXT:
            out["word"].append(f)
        elif ext in XDW_EXT:
            out["xdw"].append(f)
        else:
            out["other"].append(f)
    return out


def scan_ppap_root(root: str) -> list:
    """
    Walk root recursively.
    Every sub-folder that contains at least one supported file
    (.xlsx / .docx / .xdw) is treated as one PPAP record.
    """
    root_path = Path(root)
    if not root_path.exists():
        return []

    records = []
    for dirpath, dirnames, filenames in os.walk(root):
        dp = Path(dirpath)
        supported = [f for f in filenames if Path(f).suffix.lower() in ALL_EXT]
        if not supported:
            continue

        files  = classify_files(dp)
        parsed = parse_folder_name(dp.name)

        try:
            rel_parts = dp.relative_to(root_path).parts
            engineer  = rel_parts[0] if len(rel_parts) >= 1 else "—"
        except ValueError:
            engineer = "—"

        records.append({
            "FOLDER_NAME":  dp.name,
            "PI":           parsed["PI"],
            "PN":           parsed["PN"],
            "CUSTOMER":     parsed["CUSTOMER"],
            "ENGINEER":     engineer,
            "FOLDER_PATH":  str(dp),
            "EXCEL_COUNT":  len(files["excel"]),
            "WORD_COUNT":   len(files["word"]),
            "XDW_COUNT":    len(files["xdw"]),
            "HAS_8D":       len(files["word"]) > 0,
            "EXCEL_FILES":  files["excel"],
            "WORD_FILES":   files["word"],
            "XDW_FILES":    files["xdw"],
        })

    records.sort(key=lambda r: r["PI"], reverse=True)
    return records


@st.cache_data(ttl=60, show_spinner="Scanning PPAP folder…")
def get_index(root: str) -> list:
    return scan_ppap_root(root)


def extract_xlsx_summary(file_path: str) -> dict:
    """Extract up to 10 key-value pairs from the first sheet."""
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        result = {}
        for row in ws.iter_rows(max_row=40, max_col=10, values_only=True):
            for i, cell in enumerate(row):
                if isinstance(cell, str) and cell.strip() and i + 1 < len(row):
                    val = row[i + 1]
                    if val is not None and len(result) < 10:
                        result[cell.strip()] = val
        wb.close()
        return result
    except Exception:
        return {}


# ─────────────────────────────────────────────────────────────────────────────
# BLOCK 3 — FILE OPENERS
# ─────────────────────────────────────────────────────────────────────────────

def open_file(path: str):
    try:
        if IS_WINDOWS:
            os.startfile(path)
        elif platform.system() == "Darwin":
            subprocess.run(["open", path], check=True)
        else:
            subprocess.run(["xdg-open", path], check=True)
    except Exception as e:
        st.error(f"Cannot open file: {e}")


def open_folder(path: str):
    try:
        if IS_WINDOWS:
            subprocess.Popen(f'explorer "{path}"')
        elif platform.system() == "Darwin":
            subprocess.run(["open", path])
        else:
            subprocess.run(["xdg-open", path])
    except Exception as e:
        st.error(f"Cannot open folder: {e}")


def open_docuworks(path: str):
    dw = next((p for p in DOCUWORKS_PATHS if os.path.exists(p)), None)
    if dw:
        subprocess.Popen([dw, path])
    else:
        open_file(path)


# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("## 📦 PPAP Manager")
    st.divider()

    folder = st.text_input(
        "PPAP root folder",
        value=st.session_state.get("folder", DEFAULT_FOLDER),
        key="folder",
        help=(
            "Local Windows: C:\\Users\\...\\Desktop\\PPAP\n"
            "Streamlit Cloud: data"
        ),
    )

    debug_mode = st.toggle(
        "🔍 Debug mode", value=False,
        help="Show all files and folders the app can see"
    )

    if st.button("🔄  Re-scan folder", use_container_width=True):
        st.cache_data.clear()
        st.rerun()

    records = get_index(folder)
    df_meta = pd.DataFrame([
        {"PI": r["PI"], "PN": r["PN"],
         "CUSTOMER": r["CUSTOMER"], "HAS_8D": r["HAS_8D"]}
        for r in records
    ]) if records else pd.DataFrame()

    st.divider()
    c1, c2 = st.columns(2)
    c1.metric("PPAP records", len(records))
    c2.metric("Customers",
              df_meta["CUSTOMER"].nunique() if not df_meta.empty else 0)
    st.metric("Records with 8D docs ⚠",
              int(df_meta["HAS_8D"].sum()) if not df_meta.empty else 0)

    st.divider()
    st.caption(f"Refreshed: {datetime.now().strftime('%d/%m/%Y  %H:%M')}")
    st.caption(f"Root: `{folder}`")

    # Environment hint
    if IS_WINDOWS:
        st.caption("🖥 Running locally on Windows")
    else:
        st.caption("☁️ Running on Streamlit Cloud — put files in `data/` folder on GitHub")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

st.title("🔍  Search PPAP")

# ── Folder not found ──────────────────────────────────────────────────────────
if not Path(folder).exists():
    st.error(f"**Folder not found:** `{folder}`")

    if IS_WINDOWS:
        st.info(
            "**Running locally on Windows:**\n\n"
            "Make sure the folder exists. Your folder path is:\n"
            "```\nC:\\Users\\S2234009\\Desktop\\PPAP\n```\n"
            "Open File Explorer, go to Desktop, check if PPAP folder is there."
        )
    else:
        st.info(
            "**Running on Streamlit Cloud:**\n\n"
            "You need to upload your PPAP files to GitHub first.\n\n"
            "**Steps:**\n"
            "1. Go to your GitHub repo\n"
            "2. Click into the `data/` folder\n"
            "3. Click **Add file → Upload files**\n"
            "4. Upload your PPAP files\n"
            "5. Come back and click **Re-scan folder**\n\n"
            "Then set the folder field to `data` in the sidebar."
        )
    st.stop()

# ── Debug panel ───────────────────────────────────────────────────────────────
if debug_mode:
    st.subheader("🔍 Debug — Folder scan report")
    root_path = Path(folder)
    all_dirs  = [d for d in root_path.rglob("*") if d.is_dir()]
    all_files = [f for f in root_path.rglob("*") if f.is_file()]

    st.success(f"✅ Folder exists: `{folder}`")
    col1, col2, col3 = st.columns(3)
    col1.metric("Total sub-folders", len(all_dirs))
    col2.metric("Total files (all types)", len(all_files))
    col3.metric("PPAP records found", len(records))

    if records:
        st.dataframe(pd.DataFrame([{
            "Folder name": r["FOLDER_NAME"],
            "PI":          r["PI"],
            "PN":          r["PN"],
            "Customer":    r["CUSTOMER"],
            "Engineer":    r["ENGINEER"],
            "Excel":       r["EXCEL_COUNT"],
            "Word":        r["WORD_COUNT"],
            "XDW":         r["XDW_COUNT"],
            "8D?":         "⚠️" if r["HAS_8D"] else "✅",
        } for r in records]), use_container_width=True, hide_index=True)
    else:
        st.warning(
            "No PPAP records found. The app looks for sub-folders "
            "that contain at least one `.xlsx`, `.docx`, or `.xdw` file."
        )
        if all_dirs:
            st.markdown("**Sub-folders found:**")
            for d in sorted(all_dirs)[:20]:
                st.code(str(d.relative_to(root_path)))
        else:
            st.error("No sub-folders found at all. The folder appears to be empty.")

    st.divider()

# ── Empty guard ───────────────────────────────────────────────────────────────
if not records:
    st.warning(
        f"No PPAP records found in `{folder}`.\n\n"
        "Make sure your PPAP files (.xlsx / .docx / .xdw) are inside sub-folders "
        "following the naming rule:\n"
        "```\nPI-PN-CUSTOMER\n```\n"
        "Example: `S383A014-06104173AA-1CYNAO`\n\n"
        "💡 Turn on **Debug mode** in the sidebar to see what the app can find."
    )
    st.stop()

# ─────────────────────────────────────────────────────────────────────────────
# SEARCH BAR
# ─────────────────────────────────────────────────────────────────────────────

keyword = st.text_input(
    "Search",
    placeholder="Enter PI (e.g. S383A014), PN (e.g. 06104173AA), or Customer (e.g. 1CYNAO)…",
    label_visibility="collapsed",
)

# ── No keyword → show full list ───────────────────────────────────────────────
if not keyword:
    st.info("Enter a PI, PN, or Customer code above to search.")

    customers = ["All"] + sorted(set(r["CUSTOMER"] for r in records))
    sel       = st.selectbox("Filter by customer", customers)
    filtered  = records if sel == "All" else \
                [r for r in records if r["CUSTOMER"] == sel]

    st.dataframe(pd.DataFrame([{
        "PI":           r["PI"],
        "Part Number":  r["PN"],
        "Customer":     r["CUSTOMER"],
        "Engineer":     r["ENGINEER"],
        "Excel files":  r["EXCEL_COUNT"],
        "Word docs":    r["WORD_COUNT"],
        "DocuWorks":    r["XDW_COUNT"],
        "8D alert":     "⚠️ Yes" if r["HAS_8D"] else "✅ No",
    } for r in filtered]), use_container_width=True, hide_index=True)
    st.stop()

# ─────────────────────────────────────────────────────────────────────────────
# QUERY
# ─────────────────────────────────────────────────────────────────────────────

kw = keyword.strip().upper()
matched = [
    r for r in records
    if kw in r["PI"].upper()
    or kw in r["PN"].upper()
    or kw in r["CUSTOMER"].upper()
    or kw in r["ENGINEER"].upper()
    or kw in r["FOLDER_NAME"].upper()
]

if not matched:
    st.warning(f"No results found for **{keyword}**.")
    st.stop()

st.caption(f"Found **{len(matched)}** record(s) matching **{keyword}**")

# ─────────────────────────────────────────────────────────────────────────────
# RESULTS
# ─────────────────────────────────────────────────────────────────────────────

for rec in matched:

    # ── 8D Alert ─────────────────────────────────────────────────────────────
    if rec["HAS_8D"]:
        st.error(
            f"⚠️ **8D Alert** — {rec['WORD_COUNT']} Word document(s) found for "
            f"PI **{rec['PI']}**. Review root causes and corrective actions."
        )
    else:
        st.success(f"✅ PI **{rec['PI']}** — No 8D / Word reports found.")

    # ── Header ────────────────────────────────────────────────────────────────
    st.subheader(
        f"📦  PI: {rec['PI']}  |  PN: {rec['PN']}  |  Customer: {rec['CUSTOMER']}"
    )
    st.caption(
        f"Engineer: **{rec['ENGINEER']}**  ·  "
        f"Folder: `{rec['FOLDER_PATH']}`"
    )

    c1, c2, c3 = st.columns(3)
    c1.metric("Excel / Data files",  rec["EXCEL_COUNT"])
    c2.metric("Word / 8D reports",   rec["WORD_COUNT"])
    c3.metric("DocuWorks drawings",  rec["XDW_COUNT"])

    if IS_WINDOWS:
        if st.button("📁  Open folder in Explorer",
                     key=f"folder_{rec['FOLDER_NAME']}"):
            open_folder(rec["FOLDER_PATH"])

    st.divider()

    # ── Excel files ───────────────────────────────────────────────────────────
    if rec["EXCEL_FILES"]:
        st.markdown("**📊 Excel / Measurement files**")
        for fp in rec["EXCEL_FILES"]:
            with st.expander(f"`{fp.name}`", expanded=False):
                col_a, col_b = st.columns(2)

                if col_a.button("📊  View data",
                                key=f"view_{rec['FOLDER_NAME']}_{fp.name}"):
                    with st.spinner("Reading Excel…"):
                        summary = extract_xlsx_summary(str(fp))
                    if summary:
                        st.table(pd.DataFrame(
                            summary.items(), columns=["Parameter", "Value"]
                        ))
                    else:
                        st.info("No structured data found. Open the file directly.")

                if IS_WINDOWS:
                    if col_b.button("📂  Open file",
                                    key=f"open_{rec['FOLDER_NAME']}_{fp.name}"):
                        open_file(str(fp))
                        st.toast(f"Opening {fp.name}…")
                else:
                    try:
                        with open(str(fp), "rb") as f:
                            col_b.download_button(
                                "⤓  Download", data=f,
                                file_name=fp.name,
                                mime="application/vnd.openxmlformats-officedocument"
                                     ".spreadsheetml.sheet",
                                key=f"dl_{rec['FOLDER_NAME']}_{fp.name}",
                            )
                    except OSError:
                        col_b.warning("File not accessible.")

    # ── Word / 8D files ───────────────────────────────────────────────────────
    if rec["WORD_FILES"]:
        st.markdown("**📄 Word / 8D report files**")
        for fp in rec["WORD_FILES"]:
            with st.expander(f"`{fp.name}`  ⚠️", expanded=False):
                st.warning("8D / Word document — review root cause and corrective action.")
                if IS_WINDOWS:
                    if st.button("📖  Open document",
                                 key=f"word_{rec['FOLDER_NAME']}_{fp.name}"):
                        open_file(str(fp))
                        st.toast(f"Opening {fp.name}…")
                else:
                    try:
                        with open(str(fp), "rb") as f:
                            st.download_button(
                                "⤓  Download", data=f,
                                file_name=fp.name,
                                mime="application/vnd.openxmlformats-officedocument"
                                     ".wordprocessingml.document",
                                key=f"dldoc_{rec['FOLDER_NAME']}_{fp.name}",
                            )
                    except OSError:
                        st.warning("File not accessible.")

    # ── DocuWorks files ───────────────────────────────────────────────────────
    if rec["XDW_FILES"]:
        st.markdown("**🖼 DocuWorks files**")
        for fp in rec["XDW_FILES"]:
            with st.expander(f"`{fp.name}`", expanded=False):
                st.info("DocuWorks drawing / document")
                if IS_WINDOWS:
                    if st.button("🖥  Open in DocuWorks Viewer",
                                 key=f"dw_{rec['FOLDER_NAME']}_{fp.name}"):
                        open_docuworks(str(fp))
                        st.toast(f"Launching DocuWorks for {fp.name}…")
                else:
                    st.info(
                        "DocuWorks files can only be opened on a local Windows machine. "
                        "Path: `" + str(fp) + "`"
                    )

    st.divider()
